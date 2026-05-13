"""Local Excel file operations using openpyxl (no Azure/SharePoint needed)."""

import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from dotenv import load_dotenv

load_dotenv()

EXCEL_DIR = os.getenv("EXCEL_DIR", "./data")


def _resolve_path(filename: str) -> str:
    """Resolve full path for a filename within the Excel directory."""
    path = os.path.join(EXCEL_DIR, filename)
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")
    return path


def _get_workbook(filename: str, read_only=False):
    return load_workbook(_resolve_path(filename), read_only=read_only, data_only=True)


def _save_workbook(wb, filename: str):
    wb.save(_resolve_path(filename))


def list_files() -> list[str]:
    """Return list of all .xlsx files in the Excel directory."""
    return [f for f in os.listdir(EXCEL_DIR) if f.endswith(".xlsx")]


def list_worksheets(filename: str) -> list[str]:
    """Return list of worksheet names in the Excel file."""
    wb = _get_workbook(filename, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_range(filename: str, sheet_name: str, address: str) -> list[list]:
    """
    Read a range from a worksheet.

    Args:
        filename: Excel file name (e.g. "invoice_jan_2026.xlsx")
        sheet_name: Name of the worksheet (e.g. "Sheet1")
        address: Cell range (e.g. "A1:D10")

    Returns:
        2D list of cell values.
    """
    wb = _get_workbook(filename, read_only=True)
    ws = wb[sheet_name]
    rows = ws[address]
    # Handle single cell case
    if not isinstance(rows, tuple):
        rows = ((rows,),)
    elif rows and not isinstance(rows[0], tuple):
        rows = (rows,)
    data = []
    for row in rows:
        data.append([cell.value for cell in row])
    wb.close()
    return data


def update_range(filename: str, sheet_name: str, address: str, values: list[list]) -> dict:
    """
    Update a range in a worksheet.

    Args:
        filename: Excel file name (e.g. "invoice_jan_2026.xlsx")
        sheet_name: Name of the worksheet
        address: Cell range to update (e.g. "A1:B2")
        values: 2D list of values to write

    Returns:
        Summary dict.
    """
    wb = _get_workbook(filename)
    ws = wb[sheet_name]

    # Parse start cell from address
    if ":" in address:
        start_cell = address.split(":")[0]
    else:
        start_cell = address

    # Get starting row and column
    from openpyxl.utils.cell import coordinate_from_string
    col_str, start_row = coordinate_from_string(start_cell)
    start_col = column_index_from_string(col_str)

    for r_idx, row_data in enumerate(values):
        for c_idx, value in enumerate(row_data):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)

    _save_workbook(wb, filename)
    return {"status": "success", "range": address, "rows_written": len(values)}


def add_rows(filename: str, sheet_name: str, start_column: str, row_data: list[list]) -> dict:
    """
    Append rows to the next available row in a worksheet.

    Args:
        filename: Excel file name (e.g. "invoice_jan_2026.xlsx")
        sheet_name: Name of the worksheet
        start_column: Starting column letter (e.g. "A")
        row_data: 2D list of rows to append

    Returns:
        Summary dict.
    """
    wb = _get_workbook(filename)
    ws = wb[sheet_name]

    next_row = ws.max_row + 1
    start_col = column_index_from_string(start_column)

    for r_idx, row in enumerate(row_data):
        for c_idx, value in enumerate(row):
            ws.cell(row=next_row + r_idx, column=start_col + c_idx, value=value)

    _save_workbook(wb, filename)
    num_cols = len(row_data[0])
    end_col = get_column_letter(start_col + num_cols - 1)
    end_row = next_row + len(row_data) - 1
    address = f"{start_column}{next_row}:{end_col}{end_row}"
    return {"status": "success", "range": address, "rows_appended": len(row_data)}
