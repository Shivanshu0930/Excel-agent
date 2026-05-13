"""Streamlit UI for the Excel Agent."""

import streamlit as st
import os
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

from agent import create_agent

st.set_page_config(page_title="Excel Agent", page_icon="📊", layout="wide")
st.title("📊 Excel Agent")
st.caption("Ask me to read, update, or add data to your Excel files using natural language.")

excel_dir = os.getenv("EXCEL_DIR", "./data")

# Sidebar with file info and data viewer
with st.sidebar:
    st.header("📁 Available Files")
    if os.path.isdir(excel_dir):
        files = [f for f in os.listdir(excel_dir) if f.endswith(".xlsx")]
        for f in files:
            st.markdown(f"- `{f}`")
    else:
        st.warning("Data directory not found.")
        files = []

    st.divider()
    st.markdown("**Example prompts:**")
    st.markdown("- Show me all files")
    st.markdown("- Read first 5 rows from January invoice")
    st.markdown("- Update price of ORD-202601-0003 to 450")
    st.markdown("- Add a new order to March invoice")

# Initialize agent and chat history in session state
if "agent" not in st.session_state:
    st.session_state.agent = create_agent()

if "messages" not in st.session_state:
    st.session_state.messages = []

# Layout: chat on left, data viewer on right
chat_col, viewer_col = st.columns([3, 2])

with viewer_col:
    st.subheader("📄 Data Viewer")
    if files:
        selected_file = st.selectbox("Select file to view", files)
        filepath = os.path.join(excel_dir, selected_file)
        wb = load_workbook(filepath, read_only=True, data_only=True)
        selected_sheet = st.selectbox("Select sheet", wb.sheetnames)
        ws = wb[selected_sheet]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        wb.close()

        if data:
            df = pd.DataFrame(data[1:], columns=data[0])
            st.dataframe(df, width="stretch", height=400)
            st.caption(f"{len(df)} rows × {len(df.columns)} columns")
        else:
            st.info("Sheet is empty.")
    else:
        st.info("No Excel files found in data directory.")

with chat_col:
    st.subheader("💬 Chat")

    # Display chat history
    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

# Chat input (must be outside columns)
if prompt := st.chat_input("Ask something about your Excel files..."):
    # Show user message
    st.session_state.messages.append({"role": "user", "content": prompt})

    # Get agent response
    with st.spinner("Thinking..."):
        response = st.session_state.agent.invoke(
            {"messages": [("human", m["content"]) if m["role"] == "user" else ("ai", m["content"]) for m in st.session_state.messages]}
        )
        ai_msg = response["messages"][-1].content

    st.session_state.messages.append({"role": "assistant", "content": ai_msg})
    st.rerun()
